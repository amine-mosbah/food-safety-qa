#!/usr/bin/env python3
"""
fsqa — Food Safety Q&A Generator CLI.

A single entrypoint for the Food Safety Q&A pipeline:

  fsqa setup        # one-time: paste your OpenAI API key
  fsqa explore      # show the official dataset broken down by cluster dimension
  fsqa sample       # stratified sample N records from any cluster
  fsqa run          # run the 4-stage pipeline (Q-gen → items → A-gen → judge)
  fsqa wizard       # interactive end-to-end walkthrough

The default input is `data/food_recall_incidents.csv` (7,546 records, the official
SemEval-2025 Task 9 extended release). Every command has --input so you can swap
in any file with the same schema.

See README.md for the quickstart and docs/DESIGN.md for design rationale.
"""
from __future__ import annotations

import csv
import json
import os
import random
import re
import string
import sys
import time
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Optional
from urllib.error import HTTPError
import urllib.request

try:
    import typer
    from rich.console import Console
    from rich.panel import Panel
    from rich.progress import (
        BarColumn,
        Progress,
        SpinnerColumn,
        TextColumn,
        TimeElapsedColumn,
        TimeRemainingColumn,
    )
    from rich.prompt import Confirm, IntPrompt, Prompt
    from rich.table import Table
except ImportError:
    sys.exit(
        "Missing dependencies. Install with:\n"
        "  pip install -r requirements.txt"
    )

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing openpyxl. Install with: pip install -r requirements.txt")


# ─── Constants ───────────────────────────────────────────────────────────────
DEFAULT_INPUT = "data/food_recall_incidents.csv"
DEFAULT_API_KEY_PATH = Path.home() / ".config" / "fsqa" / "openai-key"
PROMPTS_DIR_DEFAULT = "prompts"
MODEL = "gpt-4o-mini"
OPENAI_URL = "https://api.openai.com/v1/chat/completions"
TEMPS = {"qgen": 0.7, "items": 0.0, "agen": 0.0, "judge": 0.0}
MAX_RETRIES = 3
RETRY_BACKOFF = 4

# gpt-4o-mini pricing (USD per 1M tokens, May 2026)
PRICE_IN = 0.15
PRICE_OUT = 0.60

PREFIX = {
    "allergens": "alg",
    "biological": "bio",
    "foreign bodies": "fbd",
    "fraud": "frd",
    "chemical": "chm",
    "other hazard": "oth",
    "packaging defect": "pkg",
    "organoleptic aspects": "org",
    "food additives and flavourings": "faf",
    "migration": "mig",
}

CLUSTER_DIMENSIONS = ["hazard-category", "product-category", "country", "year"]

console = Console()
app = typer.Typer(
    help="Food Safety Q&A Generator — 4-stage synthetic Q&A pipeline.",
    add_completion=False,
    no_args_is_help=True,
)


# ─── Helpers ─────────────────────────────────────────────────────────────────
def load_rows(input_path: Path) -> list[dict[str, Any]]:
    """Read the official CSV (handles the unnamed index column gracefully)."""
    if not input_path.exists():
        raise typer.BadParameter(f"Input CSV not found: {input_path}")
    with open(input_path, newline="") as f:
        return list(csv.DictReader(f))


def get_api_key(explicit: Optional[str] = None) -> str:
    """Resolve the OpenAI key in priority order:
      (1) explicit --api-key flag
      (2) the file written by `fsqa setup` (~/.config/fsqa/openai-key)
      (3) the OPENAI_API_KEY environment variable (last resort)

    The stored file is preferred over the env var because shared/managed
    environments (Docker, OCPlatform, dev containers) often inject an
    OPENAI_API_KEY that points at a *different* account than the user expects,
    silently routing requests — and bills — to the wrong place.
    """
    if explicit:
        console.print(f"[dim]Using --api-key flag (…{explicit.strip()[-6:]})[/dim]")
        return explicit.strip()
    if DEFAULT_API_KEY_PATH.exists():
        key = DEFAULT_API_KEY_PATH.read_text().strip()
        console.print(f"[dim]Using key from {DEFAULT_API_KEY_PATH} (…{key[-6:]})[/dim]")
        return key
    env = os.environ.get("OPENAI_API_KEY")
    if env:
        console.print(f"[yellow]Using OPENAI_API_KEY env var (…{env.strip()[-6:]}) — run `fsqa setup` to lock in a specific key.[/yellow]")
        return env.strip()
    raise typer.BadParameter(
        "No OpenAI API key found. Run `fsqa setup` or set OPENAI_API_KEY."
    )


def slugify(text: str) -> str:
    s = re.sub(r"[^a-zA-Z0-9]+", "-", text.lower()).strip("-")
    return s or "cluster"


# ─── Prompt loading (same convention as v0.2) ────────────────────────────────
def load_prompts(prompts_dir: Path) -> dict[str, dict[str, str]]:
    files = {
        "qgen": prompts_dir / "01_question_generation.md",
        "items": prompts_dir / "02_required_items_extraction.md",
        "agen": prompts_dir / "03_answer_generation.md",
        "judge": prompts_dir / "04_llm_as_judge.md",
    }
    out = {}
    for stage, fp in files.items():
        text = fp.read_text()
        _, _, rest = text.partition("## System Prompt")
        sys_section, _, rest2 = rest.partition("## User Prompt Template")
        usr_section, _, _ = rest2.partition("\n## ")
        out[stage] = {
            "system": _extract_code_block(sys_section).strip(),
            "user_template": _extract_code_block(usr_section).strip(),
        }
    return out


def _extract_code_block(md: str) -> str:
    lines = md.splitlines()
    in_block, body = False, []
    for line in lines:
        if line.lstrip().startswith("```"):
            if not in_block:
                in_block = True
                continue
            break
        if in_block:
            body.append(line)
    return "\n".join(body) if body else md.strip()


def call_openai(api_key: str, system: str, user: str, temperature: float) -> dict[str, Any]:
    payload = {
        "model": MODEL,
        "messages": [
            {"role": "system", "content": system},
            {"role": "user", "content": user},
        ],
        "temperature": temperature,
        "response_format": {"type": "json_object"},
    }
    body = json.dumps(payload).encode()
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    last_err = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            req = urllib.request.Request(OPENAI_URL, data=body, headers=headers, method="POST")
            with urllib.request.urlopen(req, timeout=120) as resp:
                raw = json.loads(resp.read())
            return {
                "content": raw["choices"][0]["message"]["content"],
                "usage": raw.get("usage", {}),
            }
        except HTTPError as e:
            err_body = e.read().decode(errors="replace")[:300]
            last_err = f"HTTP {e.code}: {err_body}"
            if e.code in (429, 500, 502, 503, 504):
                time.sleep(RETRY_BACKOFF * attempt)
                continue
            raise RuntimeError(last_err)
        except Exception as e:  # noqa: BLE001
            last_err = f"{type(e).__name__}: {e}"
            time.sleep(RETRY_BACKOFF * attempt)
    raise RuntimeError(f"OpenAI call failed after {MAX_RETRIES} retries: {last_err}")


def fmt(template: str, record: dict[str, Any], **extras) -> str:
    fields = {
        "record_id": record.get("record_id", ""),
        "year": record.get("year", ""),
        "month": record.get("month", ""),
        "day": record.get("day", ""),
        "country": record.get("country", ""),
        "product": record.get("product", ""),
        "product_category": record.get("product-category", ""),
        "hazard": record.get("hazard", ""),
        "hazard_category": record.get("hazard-category", ""),
        "title": record.get("title", ""),
        "text": record.get("text", ""),
    }
    fields.update(extras)
    return string.Template(template).safe_substitute(fields)


def parse_json_safely(content: str) -> dict[str, Any]:
    s = content.strip()
    if s.startswith("```"):
        s = s.split("```", 2)[1]
        if s.startswith("json\n"):
            s = s[5:]
        if s.endswith("```"):
            s = s[:-3]
    return json.loads(s)


# ─── Stratified sampler ──────────────────────────────────────────────────────
def stratified_sample(
    rows: list[dict],
    by_column: str,
    value: str,
    n: int,
    stratify_on: str = "hazard",
    top_k: int = 10,
    seed: int = 42,
) -> list[dict]:
    """Pick N records where rows[by_column] == value, stratified across stratify_on."""
    random.seed(seed)
    cluster_rows = [r for r in rows if str(r.get(by_column, "")) == str(value)]
    if not cluster_rows:
        raise typer.BadParameter(
            f"No rows where {by_column}={value!r}. Run `fsqa explore` to see options."
        )
    if n >= len(cluster_rows):
        return cluster_rows

    sub_counts = Counter(r.get(stratify_on, "") for r in cluster_rows)
    top = [v for v, _ in sub_counts.most_common(top_k)]
    per_top = max(1, n // len(top))

    sample, picked_ids = [], set()
    for v in top:
        pool = [r for r in cluster_rows if r.get(stratify_on, "") == v]
        random.shuffle(pool)
        for r in pool[:per_top]:
            sample.append(r)
            picked_ids.add(id(r))
    # Top up from the long tail
    rest = [r for r in cluster_rows if id(r) not in picked_ids]
    random.shuffle(rest)
    while len(sample) < n and rest:
        sample.append(rest.pop())
    return sample[:n]


def assign_record_ids(sample: list[dict], by_column: str, value: str) -> list[dict]:
    if by_column == "hazard-category":
        prefix = PREFIX.get(value, slugify(value)[:3])
    else:
        prefix = slugify(value)[:6]
    for i, r in enumerate(sample, start=1):
        r["record_id"] = f"{prefix}_{i:03d}"
    return sample


# ─── Pipeline stages ─────────────────────────────────────────────────────────
def stage_qgen(api_key, prompts, rec):
    r = call_openai(api_key, prompts["qgen"]["system"],
                    fmt(prompts["qgen"]["user_template"], rec), TEMPS["qgen"])
    return parse_json_safely(r["content"]), r["usage"]


def stage_items(api_key, prompts, rec, question):
    r = call_openai(api_key, prompts["items"]["system"],
                    fmt(prompts["items"]["user_template"], rec, question=question), TEMPS["items"])
    return parse_json_safely(r["content"]), r["usage"]


def stage_agen(api_key, prompts, rec, question):
    r = call_openai(api_key, prompts["agen"]["system"],
                    fmt(prompts["agen"]["user_template"], rec, question=question), TEMPS["agen"])
    return parse_json_safely(r["content"]), r["usage"]


def stage_judge(api_key, prompts, rec, question, required, optional, answer):
    r = call_openai(
        api_key, prompts["judge"]["system"],
        fmt(
            prompts["judge"]["user_template"], rec,
            question=question,
            required_items_json=json.dumps(required, ensure_ascii=False),
            optional_items_json=json.dumps(optional, ensure_ascii=False),
            candidate_answer=answer,
        ),
        TEMPS["judge"],
    )
    return parse_json_safely(r["content"]), r["usage"]


# ─── Excel writer ────────────────────────────────────────────────────────────
COLUMNS = [
    ("record_id", 14), ("hazard_category", 16), ("hazard", 28),
    ("product_category", 22), ("product", 22), ("country", 8), ("year", 8),
    ("title", 50), ("question", 60), ("justification", 60),
    ("required_items", 40), ("optional_items", 40), ("answer", 60),
    ("required_coverage", 12), ("optional_coverage", 12),
    ("overall_score", 12), ("verdict", 50), ("source_text_excerpt", 60),
]


def write_excel(out_path: Path, rows: list[dict]):
    wb = Workbook()
    ws = wb.active
    ws.title = "qa_pairs"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A3A6E")
    wrap = Alignment(wrap_text=True, vertical="top")
    ws.append([c[0] for c in COLUMNS])
    for i, (name, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=i)
        c.font = header_font
        c.fill = header_fill
        ws.column_dimensions[get_column_letter(i)].width = width
    for r in rows:
        ws.append([r.get(c[0], "") for c in COLUMNS])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = wrap
    ws.freeze_panes = "A2"
    wb.save(out_path)


# ─── Commands ────────────────────────────────────────────────────────────────
@app.command()
def setup(
    api_key: Optional[str] = typer.Option(None, "--api-key", help="Paste your OpenAI sk-... key. Omit to be prompted securely."),
):
    """Store your OpenAI API key locally (chmod 600). Used by all other commands."""
    DEFAULT_API_KEY_PATH.parent.mkdir(parents=True, exist_ok=True)
    if not api_key:
        api_key = Prompt.ask(
            "Paste your OpenAI API key", password=True, console=console
        )
    api_key = api_key.strip()
    if not api_key.startswith(("sk-", "sk_")):
        console.print("[red]That doesn't look like an OpenAI key (should start with sk-).[/red]")
        raise typer.Exit(1)
    DEFAULT_API_KEY_PATH.write_text(api_key)
    DEFAULT_API_KEY_PATH.chmod(0o600)
    console.print(f"[green]✓[/green] API key saved → {DEFAULT_API_KEY_PATH} (chmod 600)")


@app.command()
def explore(
    input: Path = typer.Option(DEFAULT_INPUT, "--input", "-i", help="CSV in SemEval schema."),
    by: str = typer.Option("hazard-category", "--by", help=f"Cluster dimension. One of: {', '.join(CLUSTER_DIMENSIONS)}"),
    top: int = typer.Option(20, "--top", help="Show top N values."),
):
    """Show the dataset broken down by a cluster dimension."""
    rows = load_rows(input)
    console.print(Panel.fit(
        f"[bold]{input.name}[/bold] · {len(rows):,} records · cluster by [bold cyan]{by}[/bold cyan]",
        border_style="cyan",
    ))
    if by not in rows[0]:
        console.print(f"[red]Column {by!r} not found. Available: {list(rows[0].keys())}[/red]")
        raise typer.Exit(1)
    counts = Counter(r.get(by, "") for r in rows)
    table = Table(show_header=True, header_style="bold magenta")
    table.add_column("#", justify="right")
    table.add_column(by)
    table.add_column("records", justify="right")
    table.add_column("share", justify="right")
    total = sum(counts.values())
    for i, (val, c) in enumerate(counts.most_common(top), start=1):
        table.add_row(str(i), str(val) or "(empty)", f"{c:,}", f"{100*c/total:.1f}%")
    console.print(table)
    if len(counts) > top:
        console.print(f"[dim]…and {len(counts) - top} more values. Use --top to see more.[/dim]")


@app.command()
def sample(
    input: Path = typer.Option(DEFAULT_INPUT, "--input", "-i"),
    by: str = typer.Option("hazard-category", "--by", help=f"Cluster dimension. {', '.join(CLUSTER_DIMENSIONS)}"),
    value: str = typer.Option(..., "--value", "-v", help="Cluster value (e.g. allergens)"),
    n: int = typer.Option(50, "--n", help="Sample size"),
    stratify: str = typer.Option("hazard", "--stratify", help="Inner field to stratify the sample on"),
    out: Path = typer.Option(None, "--out", "-o", help="Output CSV path (default: samples/<value>_<n>.csv)"),
    seed: int = typer.Option(42, "--seed"),
):
    """Stratified sample N records from a chosen cluster, write a CSV with stable record_ids."""
    rows = load_rows(input)
    if out is None:
        out = Path("samples") / f"{slugify(value)}_{n}.csv"
    out.parent.mkdir(parents=True, exist_ok=True)

    sampled = stratified_sample(rows, by, value, n, stratify_on=stratify, seed=seed)
    sampled = assign_record_ids(sampled, by, value)

    fields = ["record_id"] + [k for k in sampled[0].keys() if k != "record_id"]
    with open(out, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(sampled)

    console.print(Panel.fit(
        f"[green]✓[/green] Sampled [bold]{len(sampled)}[/bold] records "
        f"({by}={value!r}, stratified on {stratify!r})\n[dim]→ {out}[/dim]",
        border_style="green",
    ))


@app.command()
def run(
    input: Path = typer.Option(..., "--input", "-i", help="Sample CSV (from `fsqa sample`)"),
    out: Path = typer.Option("outputs/qa_dataset.xlsx", "--out", "-o", help="Output Excel"),
    raw_log: Path = typer.Option("outputs/raw_runs.jsonl", "--raw-log", help="Raw LLM-call JSONL log"),
    limit: Optional[int] = typer.Option(None, "--limit", help="Process at most N records (smoke test)"),
    prompts_dir: Path = typer.Option(PROMPTS_DIR_DEFAULT, "--prompts-dir"),
    api_key: Optional[str] = typer.Option(None, "--api-key", help="Override stored key"),
):
    """Run the 4-stage pipeline (Q-gen → items → A-gen → judge) and write an Excel."""
    key = get_api_key(api_key)
    out.parent.mkdir(parents=True, exist_ok=True)
    raw_log.parent.mkdir(parents=True, exist_ok=True)

    prompts = load_prompts(prompts_dir.resolve())
    records = load_rows(input)
    if limit:
        records = records[:limit]

    console.print(Panel.fit(
        f"[bold]Pipeline run[/bold]\n"
        f"  input:   {input}  ({len(records)} records)\n"
        f"  output:  {out}\n"
        f"  model:   {MODEL}\n"
        f"  raw log: {raw_log}",
        border_style="cyan",
    ))

    rows: list[dict] = []
    usage = {"prompt_tokens": 0, "completion_tokens": 0}
    failures = 0
    scores: list[float] = []
    req_covs: list[float] = []
    opt_covs: list[float] = []

    progress = Progress(
        SpinnerColumn(),
        TextColumn("[bold blue]{task.description}"),
        BarColumn(),
        TextColumn("{task.completed}/{task.total}"),
        TimeElapsedColumn(),
        TimeRemainingColumn(),
        console=console,
    )

    with progress, open(raw_log, "w") as raw_fp:
        task = progress.add_task("Generating Q&A pairs…", total=len(records))
        for i, rec in enumerate(records, start=1):
            rid = rec.get("record_id") or f"row_{i}"
            try:
                qjson, u1 = stage_qgen(key, prompts, rec)
                question = qjson["question"]
                justification = qjson["justification"]

                ijson, u2 = stage_items(key, prompts, rec, question)
                required = ijson.get("required_items", [])
                optional = ijson.get("optional_items", [])

                ajson, u3 = stage_agen(key, prompts, rec, question)
                answer = ajson["answer"]

                jjson, u4 = stage_judge(key, prompts, rec, question, required, optional, answer)
                req_cov = float(jjson.get("required_coverage", 0.0))
                opt_cov = float(jjson.get("optional_coverage", 0.0))
                overall = float(jjson.get("overall_score", 0.0))
                verdict = jjson.get("verdict", "")

                for u in (u1, u2, u3, u4):
                    usage["prompt_tokens"] += u.get("prompt_tokens", 0)
                    usage["completion_tokens"] += u.get("completion_tokens", 0)

                rows.append({
                    "record_id": rid,
                    "hazard_category": rec.get("hazard-category", ""),
                    "hazard": rec.get("hazard", ""),
                    "product_category": rec.get("product-category", ""),
                    "product": rec.get("product", ""),
                    "country": rec.get("country", ""),
                    "year": rec.get("year", ""),
                    "title": rec.get("title", ""),
                    "question": question,
                    "justification": justification,
                    "required_items": json.dumps(required, ensure_ascii=False),
                    "optional_items": json.dumps(optional, ensure_ascii=False),
                    "answer": answer,
                    "required_coverage": req_cov,
                    "optional_coverage": opt_cov,
                    "overall_score": overall,
                    "verdict": verdict,
                    "source_text_excerpt": (rec.get("text") or "")[:300],
                })
                req_covs.append(req_cov)
                opt_covs.append(opt_cov)
                scores.append(overall)
                raw_fp.write(json.dumps({
                    "record_id": rid, "qgen": qjson, "items": ijson,
                    "agen": ajson, "judge": jjson,
                }, ensure_ascii=False) + "\n")
                raw_fp.flush()

                cost = (usage["prompt_tokens"]/1e6)*PRICE_IN + (usage["completion_tokens"]/1e6)*PRICE_OUT
                progress.update(
                    task, advance=1,
                    description=f"[blue]Q&A {i}/{len(records)}[/blue] "
                                f"[dim]· req={req_cov:.2f} score={overall:.2f} · $"
                                f"{cost:.3f} spent[/dim]",
                )
            except Exception as e:  # noqa: BLE001
                failures += 1
                console.print(f"  [red]✗[/red] {rid}: {e}")
                raw_fp.write(json.dumps({"record_id": rid, "error": str(e)}) + "\n")
                raw_fp.flush()
                progress.update(task, advance=1)

    write_excel(out, rows)
    cost = (usage["prompt_tokens"]/1e6)*PRICE_IN + (usage["completion_tokens"]/1e6)*PRICE_OUT

    # Summary
    summary = Table(title="Run summary", show_header=True, header_style="bold green")
    summary.add_column("Metric")
    summary.add_column("Value", justify="right")
    summary.add_row("Records processed", f"{len(rows)} / {len(records)}")
    summary.add_row("Failures", str(failures))
    if scores:
        summary.add_row("Avg required_coverage", f"{sum(req_covs)/len(req_covs):.3f}")
        summary.add_row("Avg optional_coverage", f"{sum(opt_covs)/len(opt_covs):.3f}")
        summary.add_row("Avg overall_score", f"{sum(scores)/len(scores):.3f}")
        summary.add_row("% req_cov = 1.0", f"{100*sum(1 for x in req_covs if x>=1.0)/len(req_covs):.0f}%")
    summary.add_row("Tokens in / out", f"{usage['prompt_tokens']:,} / {usage['completion_tokens']:,}")
    summary.add_row("[bold]Cost (USD)[/bold]", f"[bold]${cost:.4f}[/bold]")
    console.print(summary)
    console.print(f"\n[green]✓[/green] Excel: [bold]{out}[/bold]")
    console.print(f"[green]✓[/green] Raw log: [bold]{raw_log}[/bold]\n")


@app.command()
def wizard(
    input: Path = typer.Option(DEFAULT_INPUT, "--input", "-i"),
):
    """Interactive walkthrough: explore → pick cluster → sample → run."""
    if not input.exists():
        console.print(f"[red]Input not found: {input}[/red]")
        raise typer.Exit(1)

    rows = load_rows(input)
    console.print(Panel.fit(
        f"[bold cyan]fsqa wizard[/bold cyan]\n"
        f"Dataset: [bold]{input.name}[/bold] · {len(rows):,} records",
        border_style="cyan",
    ))

    # 1. Pick cluster dimension
    console.print("\n[bold]Step 1.[/bold] Cluster dimension:")
    for i, dim in enumerate(CLUSTER_DIMENSIONS, start=1):
        console.print(f"  [cyan]{i}[/cyan]) {dim}")
    pick = IntPrompt.ask("Pick a dimension", default=1, choices=[str(i) for i in range(1, len(CLUSTER_DIMENSIONS)+1)])
    by = CLUSTER_DIMENSIONS[pick - 1]

    # 2. Pick cluster value
    counts = Counter(r.get(by, "") for r in rows).most_common()
    console.print(f"\n[bold]Step 2.[/bold] Cluster value (sorted by size):")
    table = Table(show_header=True, header_style="bold magenta")
    table.add_column("#", justify="right")
    table.add_column(by)
    table.add_column("records", justify="right")
    for i, (val, c) in enumerate(counts[:15], start=1):
        table.add_row(str(i), str(val) or "(empty)", f"{c:,}")
    console.print(table)
    pick = IntPrompt.ask("Pick a cluster", default=1, choices=[str(i) for i in range(1, min(15, len(counts))+1)])
    value, value_count = counts[pick - 1]

    # 3. Sample size
    default_n = min(50, value_count)
    n = IntPrompt.ask(
        f"\n[bold]Step 3.[/bold] Sample size (cluster has {value_count:,} records)",
        default=default_n,
    )

    # 4. Estimate cost
    est = n * 0.025
    console.print(
        f"\n[bold]Step 4.[/bold] Estimated cost: [bold]${est:.2f}[/bold] "
        f"({n} records × ~$0.025 with gpt-4o-mini)"
    )
    if not Confirm.ask("Proceed?", default=True):
        console.print("[yellow]Aborted.[/yellow]")
        raise typer.Exit()

    # 5. Run
    sample_path = Path("samples") / f"{slugify(str(value))}_{n}.csv"
    out_path = Path("outputs") / f"qa_{slugify(str(value))}_{n}.xlsx"
    log_path = Path("outputs") / f"raw_{slugify(str(value))}_{n}.jsonl"

    sample_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    sampled = stratified_sample(rows, by, str(value), n)
    sampled = assign_record_ids(sampled, by, str(value))
    fields = ["record_id"] + [k for k in sampled[0].keys() if k != "record_id"]
    with open(sample_path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(sampled)
    console.print(f"[green]✓[/green] Sample: [bold]{sample_path}[/bold]\n")

    # Delegate to run()
    run(
        input=sample_path,
        out=out_path,
        raw_log=log_path,
        limit=None,
        prompts_dir=Path(PROMPTS_DIR_DEFAULT),
        api_key=None,
    )


if __name__ == "__main__":
    app()
